import numpy as np
import os, sys
import matplotlib.pyplot as plt
import h5py
import openpyxl

sys.path.insert(0, sys.path[0]+"\\..\\"+"..\\")
from h5py_process import *

if __name__ == "__main__":
    
    mypath = os.path.split(__file__)[0]
    os.chdir(mypath)
    # os.chdir("..")

    '''
    you need to enter the folder where the vicsekData.hdf5 file is
    for example, run the code

    os.chdir(" the path of the folder ")

    you can run 
    os.chdir("..")
    to Back to the parent catalogue

    you can use 
    nowpath = os.path.abspath('.')
    print(nowpath)
    to know where you are

    '''
    # enter the path where the data is
    os.chdir("..")
    os.chdir("..")
    os.chdir("Data")
    os.chdir("Pairwise_interaction")

    
    #to run this code you must know the list of eta and wlf

    eta_list = np.arange(0, 2.01, .05)*np.pi

    wFL_list = [1.0, 1.1, 1.2, 1.3, 1.4, 1.5, 1.7, 1.9, 2.0,3.0,4.0,5.0,7.0,9.0,10.0,15,20,25,30,40,50,60,70,80,90,100] # argv[6]

    informationList = {
        # # conditional entropy
        # "cE_box":16,
        # "cE_Guassian":16,
        
        #Kernel estimation method
        "MI_kernel":16,
        "TDMI_kernel":16,
        "TE_kernel":16,
            # normalized by mutual information
        "nTDMI_kernel":16,
        "nTE_kernel":16,

        # Kozachenko-Leonenko (KL)/ Metric/ kNN Entropy Estimation/ KSG method
        "MI_KL":16,
        "TDMI_KL":16,
        "TE_KL":16,
            # normalized by mutual information
        "nTDMI_KL":16,
        "nTE_KL":16,


        # # # Ordinal mutual information between two time series

        # "MI_ordinal":16,
        # "TDMI_ordinal": 16,
        # "TE_ordinal":16,
        #     # normalized by mutual information
        # "nTDMI_ordinal": 16,
        # "nTE_ordinal":16,
        }

    wb = openpyxl.Workbook()
    
    sheet = wb.create_sheet("information", 0)
    quantity_list = []
    eta_list = np.round(eta_list,5)
    # print(eta_list)
    for key in informationList:
        quantity_list.append(key)
        quantity_index= list(quantity_list).index(key)
        sheet.cell(row=1,column=1+quantity_index*(len(eta_list)+1)).value = key + "eta(col)_wFL(row)"


        for i in range(np.shape(eta_list)[0]):
            sheet.cell(row=1,column=2+i).value = str(round(eta_list[i]/np.pi,2))
        for i in range(np.shape(wFL_list)[0]):
            sheet.cell(row=2+i,column=1).value = str(round(wFL_list[i],2))

    
    for folder in os.listdir("."): # folder is the name of each folder
        if folder[-5:] == '.xlsx' or folder[-4:] == '.txt':
            continue
        print(folder)
        os.chdir(folder)

        try:
            f = H5PY_Processor("Data_estimator"+".hdf5", 'r')
        except:
            os.chdir("..")
            continue

        number_of_particles  = f.f["number_of_particles"][:][0]
        number_of_influencers = f.f["number_of_influencers"][:][0]
        wFL = f.f["wLF"][:][0] # the difference in name is the convention
        eta = f.f["eta"][:][0]

        dataNames = f.f.keys()
        
        for key in dataNames:
            if key in informationList:
                informationList[key] = f.f[key][:][0]

        for keys in dataNames:
            if keys in informationList:
                quantity_index= list(quantity_list).index(keys)
                eta_index= list(eta_list).index(np.round(eta,5))
                wFL_index= list(wFL_list).index(wFL)

                sheet.cell(row= 2 + wFL_index, column= 2 + eta_index+quantity_index*(len(eta_list)+1)).value = informationList[keys]#和画图的格式一样
        f.close()
        os.chdir("..")

    wb.save('Entropy_estimator'+'.xlsx')




    